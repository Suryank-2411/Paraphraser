import os
import openpyxl as xl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
import time
import json
wb = xl.load_workbook("C:\\Users\\Suryank\\Downloads\\sheet1.xlsx")
sheet_names = wb.sheetnames
no_of_sheets = len(sheet_names)
for index in range(no_of_sheets):
  ws=wb.worksheets[index]
  no_of_rows = ws.max_row
  no_of_columns = ws.max_column
  for i in range(1,no_of_rows+1):
    for j in range(1,no_of_columns+1):
      v=ws.cell(row=i,column=j).value
      print("Value from excel =" + str(v))
      if(v == None or v == ' ' or v == '  ' or v == '   '):
        continue
      os.chdir('C:\\Users\\Suryank\\Downloads')
      list_dir=os.listdir('C:\\Users\\Suryank\\Downloads')#got list of all the files in the downloads folder
      time.sleep(2)
      for files in list_dir:
          if(files == 'json.txt'):
              os.remove(files)
          elif(files == 'json (1).txt'):
              os.remove(files)
      service = Service()
      options = webdriver.ChromeOptions()
      driver = webdriver.Chrome(service=service, options=options)
      time.sleep(2)
      driver.get('https://translate.googleapis.com/translate_a/single?client=gtx&sl=en&tl=zu&dt=t&q='+str(v))
      print("Our Row Number = " + str(i))
      print("Our Box Number = " + str(j))
      print("first Conversion Started")
      time.sleep(2)
      print("first Conversion Downloaded")
      open_json_text_file_one = open('C:\\Users\\Suryank\\Downloads\\json.txt')
      json_array_inside_textfile_one = json.load(open_json_text_file_one)
      # print(json_array_inside_textfile_one)
      print("First File Array Data")
      
      if(json_array_inside_textfile_one[0] == 'null' or json_array_inside_textfile_one[0] == None):
        val = ' '
      else:
        val = json_array_inside_textfile_one[0][0][0]
        open_json_text_file_one.close()
      # print(val)
      
      driver.get('https://translate.googleapis.com/translate_a/single?client=gtx&sl=zu&tl=en&dt=t&q='+str(val))
      print("Second Conversion Started")
      time.sleep(2)
      print("Second Conversion Downloaded")
      time.sleep(2)
      open_json_text_file_two = open('C:\\Users\\Suryank\\Downloads\\json (1).txt')
      json_array_two = json.load(open_json_text_file_two)
      print("Second File Array Data")
      # vall = json_array_two[0]
      if(json_array_two[0] == 'null' or json_array_two[0] == None):
        vall = ' '
      else:
        vall = json_array_two[0][0][0]
      vall = json_array_two[0][0][0]
      print(json_array_two[0][0][0])
      ws.cell(row=i,column=j).value = vall
      open_json_text_file_two.close()
      
      os.chdir("C:\\Users\\Suryank\\Downloads")
      wb.save("C:\\Users\\Suryank\\Downloads\\sheet1_"+str((sheet_names[index]))+".xlsx")
      time.sleep(2)
print("program completed")



    